import sys
if sys.platform == 'win32':
    from waitress import serve

import falcon
import pandas as pd
from openpyxl.writer.excel import save_virtual_workbook
import openpyxl as xl
from werkzeug.wrappers import Request
import io
import UsingPandas as up
import re


CONST_FCST_FILE="ForecastFile"
CONST_ACT_FILE="ActualsFile"
CONST_RATE_FILE="RatesFile"
FCST_DATE=4
ACT_DATE='Entry Date'
ACT_IS_APPRVD='Timesheet is Approved'
HTML_XL_CONTENT_TYPE='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

class RCL(object):

    def proc_analytics(self, data_in, data_out):
        #wb = xl.load_workbook(filename=file)
        ws = data_in.active
        empresa = ws['B1'].value

        count = 0
        acct_pattern = '[1-9][0-9]*\.[0-9]*\.'
        rut_pattern = '\'[0-9]*-[0-9]/[0-9]'
        acct_re = re.compile(acct_pattern)
        rut_re = re.compile(rut_pattern)
        account = ''
        if data_out is None:
            data_out = xl.Workbook()
        ws = data_out.active
        for r in ws.iter_rows(min_row=7):
            if r[0].value is not None:
                if acct_re.match(r[0].value):
                    account = r[0].value
                elif rut_re.match(r[0].value):
                    print('{0};{1};{2}'.format(empresa, account, ';'.join([str(c.value) for c in r])))
                count += 1
            if count > 20:
                break
        return data_out

    def on_post(self, req, resp):
        wz = Request(req.env)

        files = wz.files
        if len(files) > 1:
            engine = None
            if files[CONST_FCST_FILE].filename is not None and files[CONST_FCST_FILE].filename.lower().endswith(".xls"):
                #erroneous format
                resp.status = falcon.HTTP_500
            else:
                for file in files:
                    analytics = self.proc_analytics(io.BytesIO(file.stream.read()))
                    #pd.read_excel(io.BytesIO(file.stream.read()),engine=engine, header=6, converters={FCST_DATE: up.get_sunday})

                resp.set_header("Content-Disposition", "attachment; filename=forecast_actuals.xlsx")
                resp.context_type = 'application/octet-stream;'
                resp.body = save_virtual_workbook(fplusa)
                resp.status = falcon.HTTP_200
        else:
            resp.status = falcon.HTTP_500


class PMTools(object):

    """def on_get(self, req, resp):
        wb = upd.process(fcst, act)
        resp.set_header("Content-Disposition", "attachment; filename=\"workbook.xlsx\"")
        resp.context_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.data = save_virtual_workbook(wb)
        #resp.stream = save_virtual_workbook(wb)
        resp.status = falcon.HTTP_200
        #resp.body='This site is still building'
    """
    def on_post(self, req, resp):
        wz = Request(req.env)

        files = wz.files
        if len(files) > 1:
            engine = None
            if files[CONST_FCST_FILE].filename is not None and files[CONST_FCST_FILE].filename.lower().endswith(".xls"):
                engine = 'xlrd'
            fcst = pd.read_excel(io.BytesIO(files[CONST_FCST_FILE].stream.read()),engine=engine, header=6, converters={FCST_DATE: up.get_sunday})
            engine = None
            if files[CONST_ACT_FILE].filename is not None and files[CONST_ACT_FILE].filename.lower().endswith(".xls"):
                engine = 'xlrd'
            actuals = pd.read_excel(io.BytesIO(files[CONST_ACT_FILE].stream.read()),engine=engine, header=7, converters={ACT_DATE: up.get_sunday,
                                                         ACT_IS_APPRVD: lambda x: 0 if x == 'Y' else -1})
            if files[CONST_RATE_FILE].filename is not None and files[CONST_RATE_FILE].filename.lower().endswith(".xls"):
                engine = 'xlrd'
            rates = pd.read_excel(io.BytesIO(files[CONST_RATE_FILE].stream.read()),engine=engine, header=0, converters={'Actual Billing Rate': up.parse_rate})
            fplusa = up.process(fcst,actuals,rates)

            resp.set_header("Content-Disposition", "attachment; filename=forecast_actuals.xlsx")
            resp.context_type = 'application/octet-stream;'
            resp.body = save_virtual_workbook(fplusa)
            resp.status = falcon.HTTP_200
        else:
            resp.status = falcon.HTTP_500


def main(app, pm):
    serve(app, host='127.0.0.1',port='9980')

app = falcon.API()
pm = PMTools()
app.add_route('/pmtools',pm)

rcl = RCL()
app.add_route('/analytics', rcl)

if sys.platform == 'win32':
    if __name__== '__main__':
        main(app, pm)


import openpyxl as xl
import re

file = 'D:\\dev\\projects\\EERR\\data\\RCL_ANALISIS_07.2019.xlsx'

if __name__ == '__main__':
    wb = xl.load_workbook(filename=file)
    ws = wb.active
    empresa= ws['B1'].value

    count = 0
    acct_pattern = '[1-9][0-9]*\.[0-9]*\.'
    rut_pattern = '\'[0-9]*-[0-9]/[0-9]'
    acct_re = re.compile(acct_pattern)
    rut_re = re.compile(rut_pattern)
    account = ''
    for r in ws.iter_rows(min_row=7):
        if r[0].value is not None:
            if acct_re.match(r[0].value):
                account = r[0].value
            elif rut_re.match(r[0].value):
                print('{0};{1};{2}'.format(empresa,account,';'.join([str(c.value) for c in r])))
            count += 1
        if count > 20:
            break
    wb.close()